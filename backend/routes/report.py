"""Issue report route — appends reports to Excel, uploads screenshots to Supabase Storage."""

import os
import uuid
from datetime import datetime, timezone

import openpyxl
from fastapi import APIRouter, File, Form, UploadFile
from fastapi.responses import FileResponse, JSONResponse

router = APIRouter(tags=["report"])

BASE_DIR = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
ISSUES_FILE = os.path.join(BASE_DIR, "issues.xlsx")
BUCKET = "issue-screenshots"


def _supabase():
    url = os.getenv("SUPABASE_URL", "").strip()
    key = os.getenv("SUPABASE_ANON_KEY", "").strip()
    if not url or not key:
        return None
    from supabase import create_client
    return create_client(url, key)


def _upload_screenshot(data: bytes, ext: str) -> str | None:
    """Upload bytes to Supabase Storage and return the public URL, or None on failure."""
    try:
        sb = _supabase()
        if not sb:
            return None
        filename = f"{uuid.uuid4().hex}{ext}"
        content_type = "image/png" if ext in (".png", "") else "image/jpeg"
        sb.storage.from_(BUCKET).upload(
            filename,
            data,
            {"content-type": content_type, "upsert": "false"},
        )
        result = sb.storage.from_(BUCKET).get_public_url(filename)
        return result
    except Exception:
        return None


def _get_workbook():
    if os.path.exists(ISSUES_FILE):
        wb = openpyxl.load_workbook(ISSUES_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Issues"
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 65
        ws.column_dimensions["F"].width = 80
        ws.append(["#", "Timestamp (UTC)", "Reporter Email", "Page / Context", "Description", "Screenshot URL"])
    return wb, ws


@router.post("/api/report-issue")
async def report_issue(
    description: str = Form(...),
    reporter_email: str = Form(""),
    page: str = Form(""),
    screenshot: UploadFile = File(None),
):
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    screenshot_url = "—"

    if screenshot and screenshot.filename:
        data = await screenshot.read()
        ext = os.path.splitext(screenshot.filename)[-1].lower() or ".png"
        url = _upload_screenshot(data, ext)
        screenshot_url = url if url else "upload-failed"

    try:
        wb, ws = _get_workbook()
        issue_num = ws.max_row
        ws.append([issue_num, timestamp, reporter_email or "—", page or "—", description, screenshot_url])
        wb.save(ISSUES_FILE)
        return JSONResponse({"status": "ok", "message": "Issue recorded — thank you for the report!"})
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": f"Failed to save report: {str(e)}"}
        )


@router.get("/api/issues")
async def list_issues():
    if not os.path.exists(ISSUES_FILE):
        return JSONResponse({"issues": [], "total": 0})
    try:
        wb = openpyxl.load_workbook(ISSUES_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        issues = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                issues.append(dict(zip(headers, row)))
        return JSONResponse({"issues": issues, "total": len(issues)})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/issues/download")
async def download_issues():
    if not os.path.exists(ISSUES_FILE):
        return JSONResponse(status_code=404, content={"error": "No issues file found yet."})
    return FileResponse(
        ISSUES_FILE,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="aimitra-issues.xlsx",
    )
